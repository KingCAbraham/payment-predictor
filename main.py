"""
Predictor de Probabilidad de Pago
--------------------------------
Este script utiliza un modelo de Random Forest para predecir la probabilidad de que un cliente pague su deuda,
basado en características reducidas. Los resultados se guardan en un archivo Excel con formato condicional.
El modelo y el escalador se guardan en archivos .pkl para reutilización.

Requisitos: pandas, numpy, scikit-learn, matplotlib, openpyxl, joblib
Ejecutar: python main.py
"""

import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestClassifier
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import accuracy_score, classification_report
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter  # Nueva importación
import joblib

# Fecha actual
FECHA_ACTUAL = datetime(2025, 3, 6)

def cargar_datos(ruta_archivo):
    try:
        datos = pd.read_excel(ruta_archivo)
        print("Datos cargados exitosamente. Columnas disponibles:")
        print(datos.columns.tolist())
        return datos
    except FileNotFoundError:
        print(f"Error: No se encontró el archivo {ruta_archivo}")
        return None

def preprocesar_datos(datos):
    datos = datos.copy()
    datos['FECHA_ULTIMO_PAGO'] = pd.to_datetime(datos['FECHA_ULTIMO_PAGO'], errors='coerce')
    datos['IMP_ULTIMO_PAGO'] = pd.to_numeric(datos['IMP_ULTIMO_PAGO'], errors='coerce').fillna(0)

    limite_dias = FECHA_ACTUAL - timedelta(days=90)
    datos['pago'] = datos.apply(
        lambda row: 1 if (row['IMP_ULTIMO_PAGO'] > 0 and pd.notna(row['FECHA_ULTIMO_PAGO']) and row['FECHA_ULTIMO_PAGO'] >= limite_dias) else 0,
        axis=1
    )

    columnas_caracteristicas = [
        'EDAD_CLIENTE', 'DIAS_ATRASO', 'MORATORIOS', 'SALDO_TOTAL',
        'IMP_ULTIMO_PAGO', 'MONTO_PAGOS', 'ATRASO_MAXIMO'
    ]
    X = datos[columnas_caracteristicas].copy()

    for col in columnas_caracteristicas:
        X.loc[:, col] = pd.to_numeric(X[col], errors='coerce')
        if X[col].isna().any():
            print(f"Advertencia: La columna '{col}' tiene valores no numéricos. Rellenando con la media.")
            X.loc[:, col] = X[col].fillna(X[col].mean())

    escalador = StandardScaler()
    X_escalado = escalador.fit_transform(X)

    y = datos['pago']
    return X_escalado, y, escalador, columnas_caracteristicas, datos

def entrenar_modelo(X, y):
    X_entrenamiento, X_prueba, y_entrenamiento, y_prueba = train_test_split(
        X, y, test_size=0.2, random_state=42
    )

    modelo = RandomForestClassifier(n_estimators=100, random_state=42)
    modelo.fit(X_entrenamiento, y_entrenamiento)

    y_pred = modelo.predict(X_prueba)
    precision = accuracy_score(y_prueba, y_pred)
    print(f"Precisión del modelo: {precision:.2f}")
    print("Reporte de clasificación:")
    print(classification_report(y_prueba, y_pred))

    return modelo

def calcular_probabilidades_todos(modelo, escalador, X, datos):
    probabilidades = modelo.predict_proba(X)[:, 1]
    datos['PROBABILIDAD_PAGO'] = probabilidades * 100
    datos['CATEGORIA_PAGO'] = pd.cut(datos['PROBABILIDAD_PAGO'], 
                                     bins=[0, 30, 70, 100], 
                                     labels=['Bajo', 'Medio', 'Alto'], 
                                     include_lowest=True)
    return datos

def main():
    ruta_archivo = "datos_clientes.xlsx"
    datos = cargar_datos(ruta_archivo)
    if datos is None:
        return

    X, y, escalador, columnas_caracteristicas, datos = preprocesar_datos(datos)
    modelo = entrenar_modelo(X, y)

    joblib.dump(modelo, 'modelo_pago.pkl')
    joblib.dump(escalador, 'escalador_pago.pkl')
    print("Modelo y escalador guardados como 'modelo_pago.pkl' y 'escalador_pago.pkl'")

    datos_con_prob = calcular_probabilidades_todos(modelo, escalador, X, datos)
    print("Primeras 5 filas con probabilidades:")
    print(datos_con_prob[['CLIENTE_UNICO', 'NOMBRE_CTE', 'PROBABILIDAD_PAGO', 'CATEGORIA_PAGO']].head())

    total_clientes = len(datos_con_prob)
    alta_prob = len(datos_con_prob[datos_con_prob['PROBABILIDAD_PAGO'] > 70])
    prom_prob = datos_con_prob['PROBABILIDAD_PAGO'].mean()
    categorias = datos_con_prob['CATEGORIA_PAGO'].value_counts()
    
    print("\nResumen Ejecutivo:")
    print(f"Total de clientes: {total_clientes}")
    print(f"Clientes con alta probabilidad (>70%): {alta_prob} ({alta_prob/total_clientes:.2%})")
    print(f"Probabilidad promedio: {prom_prob:.2f}%")
    print("Distribución por categoría:")
    print(categorias)

    with pd.ExcelWriter("resultados_probabilidades.xlsx", engine='openpyxl') as writer:
        datos_con_prob.to_excel(writer, index=False, sheet_name='Resultados')
        worksheet = writer.sheets['Resultados']
        for col in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            worksheet.column_dimensions[col[0].column_letter].width = max_length + 2

        headers = [cell.value for cell in worksheet[1]]
        prob_col_idx = headers.index('PROBABILIDAD_PAGO') + 1  # Índice basado en 1
        prob_col_letter = get_column_letter(prob_col_idx)  # Convertir a letra válida

        red_fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
        green_fill = PatternFill(start_color='99FF99', end_color='99FF99', fill_type='solid')

        # Usar la letra de columna correcta
        range_string = f'{prob_col_letter}2:{prob_col_letter}{worksheet.max_row}'
        worksheet.conditional_formatting.add(
            range_string,
            CellIsRule(operator='lessThan', formula=['30'], fill=red_fill)
        )
        worksheet.conditional_formatting.add(
            range_string,
            CellIsRule(operator='between', formula=['30', '70'], fill=yellow_fill)
        )
        worksheet.conditional_formatting.add(
            range_string,
            CellIsRule(operator='greaterThan', formula=['70'], fill=green_fill)
        )

    print("Resultados guardados en 'resultados_probabilidades.xlsx' con formato condicional")

    importancias = modelo.feature_importances_
    fig, ax = plt.subplots(figsize=(10, 6))
    bars = ax.barh(columnas_caracteristicas, importancias, color='royalblue')
    ax.set_xlabel("Importancia", fontsize=12)
    ax.set_title("Importancia de las Características en la Predicción de Pago", fontsize=14, pad=10)
    ax.tick_params(axis='y', labelsize=10)
    
    for bar in bars:
        width = bar.get_width()
        ax.text(x=width + 0.005, y=bar.get_y() + bar.get_height()/2, s=f'{width:.3f}', 
                ha='left', va='center', fontsize=10, color='black')

    plt.tight_layout()
    plt.show()

if __name__ == "__main__":
    main()